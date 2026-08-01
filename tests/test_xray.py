"""Tests for the Excel X-Ray engine. Run with:  python -m pytest"""

import sys
from pathlib import Path

from openpyxl import Workbook

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from analyze import analyze, normalize_formula, _score, Finding  # noqa: E402
from report import render_html  # noqa: E402


def _cats(result):
    return {f.category for f in result.findings}


def test_normalize_makes_column_pattern_consistent():
    # =A1+B1 in C1 and =A2+B2 in C2 should normalize to the SAME relative pattern
    assert normalize_formula("=A1+B1", 3, 1) == normalize_formula("=A2+B2", 3, 2)
    # a different structure should normalize differently
    assert normalize_formula("=A1+B1", 3, 1) != normalize_formula("=A1*B1", 3, 1)


def test_detects_inconsistent_formula(tmp_path: Path):
    wb = Workbook()
    ws = wb.active
    # Column C: consistent =A+B pattern, except one row that multiplies (the bug)
    for r in range(1, 6):
        ws[f"A{r}"] = r
        ws[f"B{r}"] = r * 2
        ws[f"C{r}"] = f"=A{r}+B{r}"
    ws["C3"] = "=A3*B3"  # breaks the pattern
    path = tmp_path / "wb.xlsx"
    wb.save(path)

    result = analyze(path)
    incon = [f for f in result.findings if f.category == "Inconsistent formula"]
    assert len(incon) == 1
    assert incon[0].cell == "C3"


def test_total_row_not_flagged_as_inconsistent(tmp_path: Path):
    # A SUM total under a =B+C column is legitimate — must NOT be flagged.
    wb = Workbook()
    ws = wb.active
    for r in range(1, 6):
        ws[f"D{r}"] = f"=B{r}+C{r}"
    ws["D6"] = "=SUM(D1:D5)"
    path = tmp_path / "wb.xlsx"
    wb.save(path)
    result = analyze(path)
    assert not any(f.category == "Inconsistent formula" for f in result.findings)


def test_running_total_seed_not_flagged(tmp_path: Path):
    # First cell seeds a running total — a legit first-of-run exception.
    wb = Workbook()
    ws = wb.active
    ws["D1"] = "=C1"
    for r in range(2, 6):
        ws[f"D{r}"] = f"=D{r-1}+C{r}"
    path = tmp_path / "wb.xlsx"
    wb.save(path)
    result = analyze(path)
    assert not any(f.category == "Inconsistent formula" for f in result.findings)


def test_interior_inconsistency_still_flagged_as_medium(tmp_path: Path):
    wb = Workbook()
    ws = wb.active
    for r in range(1, 6):
        ws[f"D{r}"] = f"=B{r}+C{r}"
    ws["D3"] = "=B3+C3+100"  # interior, non-aggregate -> should be flagged
    path = tmp_path / "wb.xlsx"
    wb.save(path)
    result = analyze(path)
    incon = [f for f in result.findings if f.category == "Inconsistent formula"]
    assert len(incon) == 1
    assert incon[0].cell == "D3"
    assert incon[0].severity == "medium"


def test_volatile_inside_string_ignored(tmp_path: Path):
    wb = Workbook()
    ws = wb.active
    ws["A1"] = '="use OFFSET() carefully"'  # text mentioning a volatile fn
    path = tmp_path / "wb.xlsx"
    wb.save(path)
    result = analyze(path)
    assert not any(f.category == "Volatile function" for f in result.findings)


def test_detects_broken_ref_and_volatile(tmp_path: Path):
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "=#REF!+1"
    ws["A2"] = "=TODAY()"
    ws["A3"] = "=INDIRECT(\"A1\")"
    path = tmp_path / "wb.xlsx"
    wb.save(path)

    result = analyze(path)
    cats = _cats(result)
    assert "Broken reference" in cats
    assert "Volatile function" in cats


def test_detects_hidden_sheet(tmp_path: Path):
    wb = Workbook()
    wb.active.title = "Visible"
    hidden = wb.create_sheet("Secret")
    hidden.sheet_state = "hidden"
    path = tmp_path / "wb.xlsx"
    wb.save(path)

    result = analyze(path)
    assert "Hidden sheet" in _cats(result)


def test_clean_workbook_scores_high(tmp_path: Path):
    wb = Workbook()
    ws = wb.active
    for r in range(1, 6):
        ws[f"A{r}"] = r
        ws[f"B{r}"] = f"=A{r}*2"
    path = tmp_path / "clean.xlsx"
    wb.save(path)

    result = analyze(path)
    assert result.health >= 90
    assert result.grade == "A"


def test_score_penalizes_errors():
    findings = [Finding("high", "Formula error", "S", "A1", "x") for _ in range(3)]
    health, grade, _ = _score(findings)
    assert health < 100
    assert grade != "A" or health < 100


def test_report_renders_html(tmp_path: Path):
    wb = Workbook()
    wb.active["A1"] = "=TODAY()"
    path = tmp_path / "wb.xlsx"
    wb.save(path)
    result = analyze(path)
    html_out = render_html(result, generated="test")
    # embeddable fragment: scoped wrapper, no document scaffolding, no header/footer
    assert '<div class="xray">' in html_out
    assert "<!doctype" not in html_out.lower()
    assert "<body" not in html_out.lower()
    assert "HEALTH / 100" in html_out
    assert str(result.health) in html_out
