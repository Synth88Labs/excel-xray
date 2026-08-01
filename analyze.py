"""
analyze.py — The Excel X-Ray analysis engine.

Reads a workbook and produces a structured risk analysis: formula errors,
inconsistent formulas (a cell that breaks its column's pattern), external
links, volatile functions, and hidden sheets — plus a Health Score.

This module has no CLI and no HTML; it just returns data structures, so it's
easy to test and to reuse behind a web UI.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook

# --- Excel error values that can appear in cached results ---
ERROR_VALUES = {
    "#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A",
    "#NULL!", "#NUM!", "#SPILL!", "#CALC!",
}

# --- Volatile functions: recalc constantly / can be unstable or slow ---
VOLATILE_FUNCS = ["NOW", "TODAY", "RAND", "RANDBETWEEN", "OFFSET", "INDIRECT", "INFO", "CELL"]

# Matches an A1-style cell reference, avoiding function names (LOG10) and table
# columns (Table[Col]) via boundary look-arounds.
_CELL_RE = re.compile(
    r"(?<![A-Za-z0-9_\]])(\$?)([A-Za-z]{1,3})(\$?)([1-9]\d{0,6})(?![A-Za-z0-9_(])"
)
_EXTERNAL_RE = re.compile(r"\[(?:\d+|[^\]]*\.xl[a-z]*)\]", re.IGNORECASE)

# Aggregate functions — a formula using one is usually a legitimate subtotal/total,
# not a bug, so it should not be flagged as an inconsistent formula.
_AGG_RE = re.compile(
    r"\b(SUM|SUMIF|SUMIFS|AVERAGE|AVERAGEIF|AVERAGEIFS|SUBTOTAL|COUNT|COUNTA|"
    r"COUNTIF|COUNTIFS|MIN|MAX|PRODUCT|MEDIAN|AGGREGATE)\s*\(", re.IGNORECASE)
_STRING_RE = re.compile(r'"[^"]*"')


@dataclass
class Finding:
    severity: str      # "high" | "medium" | "low" | "info"
    category: str
    sheet: str
    cell: str          # e.g. "C7" or "-"
    message: str
    detail: str = ""


@dataclass
class SheetInfo:
    name: str
    state: str
    rows: int
    cols: int
    formulas: int


@dataclass
class AnalysisResult:
    path: str
    sheets: list[SheetInfo] = field(default_factory=list)
    findings: list[Finding] = field(default_factory=list)
    n_sheets: int = 0
    n_cells: int = 0
    n_formulas: int = 0
    counts: dict[str, int] = field(default_factory=dict)
    health: int = 100
    grade: str = "A"
    risk_label: str = "Low risk"


# ---------- helpers ----------

def _col_to_num(letters: str) -> int:
    n = 0
    for ch in letters.upper():
        n = n * 26 + (ord(ch) - 64)
    return n


def _is_formula(value: object) -> bool:
    return isinstance(value, str) and value.startswith("=")


def normalize_formula(formula: str, col: int, row: int) -> str:
    """Convert an A1 formula into a relative R1C1-style pattern for comparison."""
    def repl(m: re.Match) -> str:
        abs_c, letters, abs_r, digits = m.groups()
        rc, rr = _col_to_num(letters), int(digits)
        cpart = f"C[${rc}]" if abs_c else f"C[{rc - col:+d}]"
        rpart = f"R[${rr}]" if abs_r else f"R[{rr - row:+d}]"
        return rpart + cpart
    return _CELL_RE.sub(repl, formula.upper())


def _volatile_in(formula: str) -> list[str]:
    # Ignore volatile keywords that appear inside string literals ("...OFFSET()...").
    up = _STRING_RE.sub("", formula).upper()
    return [f for f in VOLATILE_FUNCS if f + "(" in up]


# ---------- the analysis ----------

def analyze(path: str | Path) -> AnalysisResult:
    path = Path(path)
    wb = load_workbook(path, data_only=False)
    try:
        wb_v = load_workbook(path, data_only=True)
    except Exception:  # noqa: BLE001
        wb_v = None

    result = AnalysisResult(path=str(path))
    findings: list[Finding] = []

    for ws in wb.worksheets:
        vs = wb_v[ws.title] if wb_v is not None and ws.title in wb_v.sheetnames else None
        n_formulas = 0
        max_row = ws.max_row or 0
        max_col = ws.max_column or 0

        # column -> {row: normalized_pattern}
        col_patterns: dict[int, dict[int, str]] = {}
        # column -> {row: original A1 formula}
        col_formulas: dict[int, dict[int, str]] = {}

        for row in ws.iter_rows():
            for cell in row:
                val = cell.value
                if val is None:
                    continue
                result.n_cells += 1
                coord = cell.coordinate

                if _is_formula(val):
                    n_formulas += 1
                    result.n_formulas += 1
                    formula = str(val)

                    # broken reference
                    if "#REF!" in formula:
                        findings.append(Finding(
                            "high", "Broken reference", ws.title, coord,
                            "Formula contains #REF! — a deleted cell/range it can no longer find.",
                            formula[:120]))

                    # external links
                    if _EXTERNAL_RE.search(formula):
                        findings.append(Finding(
                            "medium", "External link", ws.title, coord,
                            "Formula links to another workbook — breaks silently if that file moves.",
                            formula[:120]))

                    # volatile functions
                    vol = _volatile_in(formula)
                    if vol:
                        findings.append(Finding(
                            "low", "Volatile function", ws.title, coord,
                            f"Uses {', '.join(vol)} — recalculates constantly; can be slow or unstable.",
                            formula[:120]))

                    # record for consistency analysis
                    c = cell.column
                    col_patterns.setdefault(c, {})[cell.row] = normalize_formula(formula, c, cell.row)
                    col_formulas.setdefault(c, {})[cell.row] = formula

                # cached error value
                if vs is not None:
                    cv = vs[coord].value
                    if isinstance(cv, str) and cv in ERROR_VALUES:
                        findings.append(Finding(
                            "high", "Formula error", ws.title, coord,
                            f"Cell evaluates to {cv}.", ""))

        # inconsistent formulas per column (within contiguous runs)
        for c, rowmap in col_patterns.items():
            rows_sorted = sorted(rowmap)
            for run in _contiguous_runs(rows_sorted):
                if len(run) < 3:
                    continue
                pats = [rowmap[r] for r in run]
                majority, maj_count = _majority(pats)
                if maj_count <= len(run) / 2:
                    continue  # no clear pattern
                for i, r in enumerate(run):
                    if rowmap[r] == majority:
                        continue
                    # Skip the common *legitimate* exceptions to cut false positives:
                    #  - the first or last cell of a run is usually a seed or a total
                    #  - a formula using an aggregate (SUM/AVERAGE…) is usually a subtotal
                    if i == 0 or i == len(run) - 1:
                        continue
                    if _AGG_RE.search(col_formulas[c][r]):
                        continue
                    coord = f"{_num_to_col(c)}{r}"
                    findings.append(Finding(
                        "medium", "Inconsistent formula", ws.title, coord,
                        "This formula differs from the column's pattern — worth a review. "
                        "It may be intentional (e.g. a subtotal or running total).",
                        col_formulas[c][r][:120]))

        result.sheets.append(SheetInfo(
            name=ws.title, state=ws.sheet_state,
            rows=max_row, cols=max_col, formulas=n_formulas))

        if ws.sheet_state in ("hidden", "veryHidden"):
            findings.append(Finding(
                "info" if ws.sheet_state == "hidden" else "low",
                "Hidden sheet", ws.title, "-",
                f"Sheet is {ws.sheet_state} — hidden logic is easy to forget and often risky.", ""))

    result.n_sheets = len(result.sheets)
    result.findings = _rank(findings)
    result.counts = _count_by_category(findings)
    result.health, result.grade, result.risk_label = _score(findings)
    return result


def _num_to_col(n: int) -> str:
    s = ""
    while n > 0:
        n, rem = divmod(n - 1, 26)
        s = chr(65 + rem) + s
    return s


def _contiguous_runs(rows: list[int]) -> list[list[int]]:
    runs, cur = [], []
    for r in rows:
        if cur and r == cur[-1] + 1:
            cur.append(r)
        else:
            if cur:
                runs.append(cur)
            cur = [r]
    if cur:
        runs.append(cur)
    return runs


def _majority(items: list[str]) -> tuple[str, int]:
    counts: dict[str, int] = {}
    for it in items:
        counts[it] = counts.get(it, 0) + 1
    best = max(counts, key=lambda k: counts[k])
    return best, counts[best]


_SEV_ORDER = {"high": 0, "medium": 1, "low": 2, "info": 3}


def _rank(findings: list[Finding]) -> list[Finding]:
    return sorted(findings, key=lambda f: (_SEV_ORDER.get(f.severity, 9), f.category, f.sheet))


def _count_by_category(findings: list[Finding]) -> dict[str, int]:
    counts: dict[str, int] = {}
    for f in findings:
        counts[f.category] = counts.get(f.category, 0) + 1
    return counts


def _score(findings: list[Finding]) -> tuple[int, str, str]:
    penalties = {
        "Formula error": (10, 40),
        "Broken reference": (10, 40),
        "Inconsistent formula": (5, 25),
        "External link": (3, 15),
        "Volatile function": (1, 10),
        "Hidden sheet": (2, 10),
    }
    per_cat: dict[str, int] = {}
    for f in findings:
        per_cat[f.category] = per_cat.get(f.category, 0) + 1
    health = 100
    for cat, count in per_cat.items():
        each, cap = penalties.get(cat, (2, 10))
        health -= min(each * count, cap)
    health = max(0, min(100, health))
    if health >= 90:
        return health, "A", "Low risk"
    if health >= 75:
        return health, "B", "Low-moderate risk"
    if health >= 60:
        return health, "C", "Moderate risk"
    if health >= 40:
        return health, "D", "High risk"
    return health, "F", "Critical risk"
